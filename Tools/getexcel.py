import openpyxl


def get_exl(file,Sheet):
    exl = openpyxl.load_workbook(file)
    table = exl[Sheet]
    max_rows = table.max_row
    max_column = table.max_column
    # print(max_rows,max_column)
    data = []
    for row in range(1, max_rows):
        rowdata = []
        for column in range(3, max_column-1):
            rowdata.append(table.cell(row+1, column+1).value)
        data.append(rowdata)
    return data


if __name__ == '__main__':
    run = get_exl('../TestData/data/testdata.xlsx','查询终端设备')
    print(run)
